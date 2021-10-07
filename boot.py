import asyncio
import discord
import openpyxl

client = discord.Client()

token = "TOKEN HERE"

@client.event

async def on_ready():
    print("Logged in as ")
    print(client.user.name)
    print(client.user.id)
    print("================")
    file = openpyxl.load_workbook("ST.xlsx")
    sheet = file.active
    sheet["A" + str(1)].value = "대기중"
    sheet["A" + str(2)].value = 0
    sheet["A" + str(3)].value = 0
    sheet["D" + str(1)].value = 0
    sheet["D" + str(2)].value = 0
    sheet["D" + str(3)].value = 0
    
    file.save("ST.xlsx")
    
    
    await client.change_presence(game=discord.Game(name="반갑습니다 :D", type=1))

@client.event

# ST -  대기중 / 모드설정 / 인원모집 / 기본게임 / 대결접수 / 게임시작 / 특별게임 / 결과출력
# ST - A[1.현재상태 / 2.인원수 / 3.완료인원 / 4.시작체널] B[인원이름] C[가위 바위 보 선택] D[결과 집계] E[등수집계]
async def on_message(message):
    if message.author.bot:
        return None

    bot_st = 0

    id = message.author.id
    me = await client.get_user_info(id)
    channel = message.channel
    
    #await client.send_message(channel, "<@"+id+">님이 \""+message.content+"\"라고 말하였습니다.")
    '''
    for i in range(2,100):
    #if sheet["A" + str(i)].value == "-":
    sheet["A" + str(i)].value = learn[1]
    sheet["B" + str(i)].value = learn[2]
    break
    '''
    file = openpyxl.load_workbook("ST.xlsx")
    sheet = file.active
    player = message.content.split(" ")
    
    if sheet["A" + str(1)].value == "대기중" and message.content.startswith('!게임시작'):    
        sheet["A" + str(1)].value = "모드설정"
        sheet["A" + str(2)].value = 0
        sheet["A" + str(4)].value = channel
        await client.send_message(channel,'게임모드를 설정합니다.')
        
    elif sheet["A" + str(1)].value != "대기중" and message.content.startswith('!게임시작'):    
        await client.send_message(channel,'게임이 진행중입니다.')
        
    elif sheet["A" + str(1)].value == "모드설정":
        if message.content.startswith('!가위바위보'):
            sheet["A" + str(1)].value = "인원모집"
            await client.send_message(channel,'인원을 모집합니다.')
    #참여 인원 모집(flag 를 이용한 중복 거르기)
    elif sheet["A" + str(1)].value == "인원모집":
        if message.content.startswith('!참여'):
            flag = 0
            playern = sheet["A" + str(2)].value
            for i in range(1,playern+1):
                if sheet["B" + str(i)].value == id:
                    flag = 1                    
            if flag == 0:
                sheet["A" + str(2)].value += 1
                #sheet.Cell(1,1).value
                sheet["B" + str(sheet["A" + str(2)].value)] = id;
                sheet["C" + str(sheet["A" + str(2)].value)] = 0
                #await client.send_message(me,"참여자가 아닙니다.")
        elif message.content.startswith('!설정완료'):
            if sheet["A" + str(2)].value <= 1:
                await client.send_message(channel, str(sheet["A" + str(2)].value) + '명은 게임진행이 불가능 합니다.')
            else:
                sheet["A" + str(1)].value = "기본게임"
                await client.send_message(channel, str(sheet["A" + str(2)].value) + '명이 게임을 시작합니다.')
                sheet["A" + str(3)].value  = 0
                for i in range(1,sheet["A" + str(2)].value + 1):
                    sheet["C" + str(i)].value = 0 
                    me = await client.get_user_info(sheet["B" + str(i)].value)
                    await client.send_message(me,"묵찌빠 선택")

    elif sheet["A" + str(1)].value == "기본게임":
        flag = 0
        for i in range(1,sheet["A" + str(2)].value + 1):
            if sheet["B" + str(i)].value == id:
                flag = i
                break
                
        if message.content.startswith('!묵') or message.content.startswith('!주먹'):
            if flag == 0:
                await client.send_message(me,"참여자가 아닙니다.")
            elif sheet["C" + str(flag)].value != 0:
                    await client.send_message(me,"이미 선택하셨습니다.")
            else:
                sheet["C" + str(flag)].value = 1
                sheet["A" + str(3)].value += 1
        elif message.content.startswith('!찌') or message.content.startswith('!가위'):
            if flag == 0:
                await client.send_message(me,"참여자가 아닙니다.")
            elif sheet["C" + str(flag)].value != 0:
                await client.send_message(me,"이미 선택하셨습니다.")
            else:
                    sheet["C" + str(flag)].value = 2
                    sheet["A" + str(3)].value += 1
        elif message.content.startswith('!빠') or message.content.startswith('!보'):
            if flag == 0:
                await client.send_message(me,"참여자가 아닙니다.")
            elif sheet["C" + str(flag)].value != 0:
                await client.send_message(me,"이미 선택하셨습니다.")
            else:
                sheet["C" + str(flag)].value = 3
                sheet["A" + str(3)].value += 1
        elif message.content.startswith('!미완료'):
            for i in range(1,sheet["A" + str(2)].value + 1):
                if sheet["C" + str(i)].value == 0:
                    await client.send_message(channel,"<@"+id+">님이 미참여하였습니다.")

        file.save("ST.xlsx")
                    
        if sheet["A" + str(2)].value == sheet["A" + str(3)].value:
            #win - 나온 손모양 카운 / win_num - 이긴 손모양 / win_cnt - 이긴사람 카운트
            win = 0
            win_num = 0;
            for i in range(1,sheet["A" + str(2)].value + 1):
                sheet["D" + str(sheet["C" + str(i)].value)].value += 1
            for i in range(1,4):
                if sheet["D" + str(i)] != 0:
                    win += 1;
                else:
                    win_num = i % 3 + 1
            if win != 2:
                for i in range(1,sheet["A" + str(2)].value + 1):
                    sheet["C" + str(i)].value = 0
                    await client.send_message(sheet["A" + str(4)].value,'비겨서 게임을 다시 시작합니다.')
            elif wine == 2 and sheet["D" + str(win_num)].value == 1:
                for i in range(1,sheet["A" + str(2)].value + 1):
                    if sheet["C" + str(i)].value == win_num:
                        id = sheet["B" + str(i)].value
                        await client.send_message(channel,"<@"+id+">님이 최종 승리하였습니다.")
                        sheet["A" + str(1)].value = "대기중"
                        break
            else:
                win_cnt = 1;
                for i in range(1,sheet["A" + str(2)].value + 1):
                    if sheet["C" + str(i)] == win_num:
                        await client.send_message(channel,"<@"+id+">님은 승리하여 게임을 계속 진행합니다.")
                        me = await client.get_user_info(sheet["B" + str(i)].value)
                        await client.send_message(me,"묵찌빠 선택")
                        sheet["B" + str(win_cnt)] = sheet["B" + str(i)]
                        win_cnt +=1;
                sheet["A" + str(2)] = sheet["D" + str(win_num)]
                
            sheet["A" + str(3)].value = 0
            sheet["D" + str(1)].value = 0
            sheet["D" + str(2)].value = 0
            sheet["D" + str(3)].value = 0
                
    
            
                    
        
        
        
                
    
    file.save("ST.xlsx")
    
    
client.run(token)
                            

'''
me=await client.get_user_info(id)

                    des='암호화 된 번호입니다.\n\n'
                    for i in range(1,n_vote+1):
                        des=des+'후보 <'+str(sheet_vote["A"+str(i+1)].value)+'> ---------- '+str(sheet_private.cell(n_parti,i).value)+'번\n\n'
                    embed=discord.Embed(title="암호화 된 번호 목록",description=des, color=0x00ffff)
                    await client.send_message(me,embed=embed)
                    '''
